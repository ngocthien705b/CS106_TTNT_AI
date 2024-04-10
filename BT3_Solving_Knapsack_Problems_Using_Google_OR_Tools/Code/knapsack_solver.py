from ortools.algorithms.python import knapsack_solver
import time
import os
import random as rd
from openpyxl import load_workbook

import pandas as pd
from pandas import DataFrame

def solve(test_file_path, time_limit = -1):
    # Input: path to input file, time_limit
    # Output:
        # test_file_path
        # problem_size
        # capacities[0]: max weight
        # computed_value
        # total_weight
        # packed_items
        # packed_values
        # packed_weights
        # run_time


    # Create the solver class.
    solver = knapsack_solver.KnapsackSolver(
        knapsack_solver.SolverType.KNAPSACK_MULTIDIMENSION_BRANCH_AND_BOUND_SOLVER,
        "KnapsackExample",
    )
    # The option KNAPSACK_MULTIDIMENSION_BRANCH_AND_BOUND_SOLVER tells the solver to use the branch and bound algorithm to solve the problem.

    values = []    # A vector containing the weights of the items
    weights = [[]]   # A vector containing the values of the items.
    capacities = []  # A vector with just one entry, the capacity of the knapsack.

    # read input from file
    with open (test_file_path, "r") as f:
        test_case = f.readlines()

        problem_size = int(test_case[1])   # n: số vật thể
        c = int(test_case[2])   # c: trọng lượng tối đa
        
        for i in range(problem_size):
            vi, wi = map(int, test_case[4+i].split() ) # vi: giá trị của vật thứ i; wi: trọng lượng của vật thứ i
            values.append(vi)
            weights[0].append(wi)
        # end for i in range(n)

        capacities.append(c)
        f.close()

    time_start = time.time()

    # Create solver object
    solver.init(values, weights, capacities)

    # set time limit
    if time_limit != -1:
        solver.set_time_limit(time_limit)

    # The total value of the optimal solution is computed_value, which is the same as the total weight in this case.
    computed_value = solver.solve()

    time_end = time.time()
    run_time = time_end - time_start

    # The program then gets the indices of the packed items in the solution
    packed_items = []   # list of the optimal packed items
    packed_weights = []     # weights of the packed items
    total_weight = 0

    for i in range(len(values)):
        #  'solver.BestSolutionContains(x)' returns 'TRUE' if the item x is included in the solution
        if solver.best_solution_contains(i):
            packed_items.append(i)
            packed_weights.append(weights[0][i])
            total_weight += weights[0][i]
        # end if solver.best_solution_contains(i)

    # end for i in range(len(values))

    packed_values = [values[i] for i in packed_items]

    return test_file_path, problem_size, capacities[0], computed_value, total_weight, packed_items, packed_values, packed_weights, run_time
# end def solve(test_file_path)

def main():
    rd.seed(42)

    # Defines the output Excel file name ("Bai_tap_3\Result.xlsx").
    file_result = "Bai_tap_3\Result.xlsx"
    
    # Lists the column names for the results dataframe.
    solution_cols = ["File path", "Problem size", "Max weight", "Total value", "Total weight", "Packed items", "Packed values", "Packed weights", "Runtime"]

    # Specifies the folder containing test cases ("testcase\kplib-master").
    test_folder = "testcase\kplib-master"

    for test_type in os.listdir(test_folder)[:1]:
        
        l_results = []

        test_path = os.path.join(test_folder, test_type)

        print("Start solving ", test_path)

        for test_size in os.listdir( test_path ):           
            
            test_limit = "R01000"

            # Constructs the path to the specific test directory.
            test_path = os.path.join(test_folder, test_type, test_size, test_limit)
            
            # Chooses a random test file from the directory.
            test_file = rd.choice( os.listdir( test_path ) )

            # Updates the test path to include the chosen file.
            test_path = os.path.join(test_folder, test_type, test_size, test_limit, test_file)
            
            # sovle teat case
            solution = solve(test_path)
            l_results.append(solution)

            print("Solved ", solution[0])

            # for i in range( len(solution) ):
            #     print(f'{solution_cols[i]}: {solution[i]}')
            # return

            # end for i in range( len(result) )

        # end for test_size in os.listdir( test_path )
        print("Done\n")
    
        # Creating Results DataFrame
        df_result = DataFrame(data = l_results, columns = solution_cols)
        
        # print(df_result.info())
        # return

        # Delete result from previous run
        if (True):
            # Load the workbook using openpyxl
            wb = load_workbook(filename=file_result)

            if test_type in wb.sheetnames:
                print("Delete result from previous run")

                # Access the sheet by name
                sheet = wb[test_type]

                # Delete the sheet
                wb.remove(sheet)

                # Save the modified workbook
                wb.save(filename=file_result)

            # end if test_type in wb.sheetnames

        # end if (True)

        # Saving Results to Excel file
        print("Saving Results to Excel file")
        with pd.ExcelWriter(path = file_result, mode = 'a') as writer:
            df_result.to_excel(excel_writer = writer, sheet_name = test_type)
        print("Done\n")
        # return

    # end for test_type in os.listdir(test_folder)

# end def main()

if __name__ == "__main__":
    main()